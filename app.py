"""Streamlit dashboard for the semiconductor vendor ESG AHP-TOPSIS model.

The app reads the existing workbook without changing it. TOPSIS is recalculated
in Python so that session-only score edits and newly uploaded workbooks are
reflected immediately, even when Excel formula caches have not been refreshed.
"""

from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import load_workbook


st.set_page_config(
    page_title="ESG Supplier Sustainability Dashboard",
    page_icon="🌿",
    layout="wide",
    initial_sidebar_state="expanded",
)

APP_DIR = Path(__file__).resolve().parent
REQUIRED_SHEETS = {
    "Level1_AHP",
    "E_AHP",
    "S_AHP",
    "G_AHP",
    "Rubric",
    "TOPSIS",
}
EVIDENCE_SHEET_NAMES = {"evidence_register", "evidence register", "evidence"}
EVIDENCE_COLUMNS = [
    "Supplier",
    "Code",
    "Indicator",
    "Evidence Type",
    "Source / URL",
    "Reporting Year",
    "Evidence Summary",
    "Score",
    "Verification Status",
    "Reviewer Notes",
    "Date Accessed",
]
PILLAR_COLORS = {
    "Environment": "#159A74",
    "Social": "#4876D1",
    "Governance": "#7857B8",
}
STATUS_COLORS = {"Complete": "#159A74", "Incomplete": "#D8952C"}


st.markdown(
    """
    <style>
      .block-container {padding-top: 1.6rem; padding-bottom: 2.5rem;}
      [data-testid="stMetric"] {
        background: linear-gradient(135deg, rgba(21,154,116,.10), rgba(72,118,209,.06));
        border: 1px solid rgba(120,130,150,.18); border-radius: 14px; padding: 14px 16px;
      }
      [data-testid="stMetricLabel"] {font-weight: 650;}
      .source-note {color: #687386; font-size: .86rem;}
    </style>
    """,
    unsafe_allow_html=True,
)


def _find_default_workbook() -> Path | None:
    """Find the model workbook in common local and deployment locations."""
    candidates: list[Path] = []
    configured = os.getenv("ESG_WORKBOOK_PATH")
    if configured:
        candidates.append(Path(configured).expanduser())
    candidates.extend(
        [
            APP_DIR / "AHP-TOPSIS Updated.xlsx",
            APP_DIR.parent / "esg-model-update" / "AHP-TOPSIS Updated.xlsx",
            Path("/mnt/data/AHP-TOPSIS Updated.xlsx"),
        ]
    )
    return next((path for path in candidates if path.is_file()), None)


def _locate_header(ws: Any, required: set[str]) -> int:
    """Return the first row containing all required header values."""
    for row in range(1, ws.max_row + 1):
        values = {
            str(ws.cell(row, col).value).strip()
            for col in range(1, ws.max_column + 1)
            if ws.cell(row, col).value is not None
        }
        if required.issubset(values):
            return row
    raise ValueError(f"Could not find headers {sorted(required)} in sheet '{ws.title}'.")


def _cell_below_label(ws: Any, label: str) -> Any:
    """Return the value immediately to the right of a row label."""
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value).strip() == label:
            return ws.cell(row, 2).value
    return None


def _read_rectangular_table(ws: Any, header_row: int) -> pd.DataFrame:
    headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
    last_header = max((idx for idx, value in enumerate(headers) if value is not None), default=-1)
    if last_header < 0:
        return pd.DataFrame()
    headers = [str(value).strip() if value is not None else f"Column {idx + 1}" for idx, value in enumerate(headers[: last_header + 1])]
    rows: list[list[Any]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, last_header + 2)]
        if not any(value is not None for value in values):
            if rows:
                break
            continue
        rows.append(values)
    return pd.DataFrame(rows, columns=headers)


def _normalize_evidence_columns(raw: pd.DataFrame) -> pd.DataFrame:
    """Map common Evidence Register headings into the stable app interface."""
    if raw.empty:
        return pd.DataFrame(columns=EVIDENCE_COLUMNS)

    aliases = {
        "supplier": "Supplier",
        "vendor": "Supplier",
        "supplier name": "Supplier",
        "code": "Code",
        "indicator code": "Code",
        "criterion code": "Code",
        "indicator": "Indicator",
        "criterion": "Indicator",
        "evidence type": "Evidence Type",
        "document type": "Evidence Type",
        "source": "Source / URL",
        "url": "Source / URL",
        "source / url": "Source / URL",
        "source url": "Source / URL",
        "reporting year": "Reporting Year",
        "year": "Reporting Year",
        "evidence summary": "Evidence Summary",
        "summary": "Evidence Summary",
        "evidence": "Evidence Summary",
        "score": "Score",
        "proposed score": "Score",
        "final score": "Score",
        "verification status": "Verification Status",
        "status": "Verification Status",
        "reviewer notes": "Reviewer Notes",
        "notes": "Reviewer Notes",
        "date accessed": "Date Accessed",
        "access date": "Date Accessed",
    }
    renamed: dict[Any, str] = {}
    for column in raw.columns:
        key = str(column).strip().lower()
        renamed[column] = aliases.get(key, str(column).strip())
    evidence = raw.rename(columns=renamed).copy()
    evidence = evidence.loc[:, ~evidence.columns.duplicated()]
    for column in EVIDENCE_COLUMNS:
        if column not in evidence.columns:
            evidence[column] = pd.NA
    evidence = evidence[EVIDENCE_COLUMNS]
    evidence = evidence.dropna(how="all")
    evidence["Score"] = pd.to_numeric(evidence["Score"], errors="coerce")
    return evidence.reset_index(drop=True)


def _parse_evidence_sheet(wb: Any) -> tuple[pd.DataFrame, str | None]:
    match = next((name for name in wb.sheetnames if name.strip().lower() in EVIDENCE_SHEET_NAMES), None)
    if match is None:
        return pd.DataFrame(columns=EVIDENCE_COLUMNS), None
    ws = wb[match]
    header_row = None
    for row in range(1, min(ws.max_row, 30) + 1):
        values = [str(ws.cell(row, col).value).strip().lower() for col in range(1, ws.max_column + 1) if ws.cell(row, col).value is not None]
        if ("supplier" in values or "vendor" in values) and ("code" in values or "indicator code" in values or "indicator" in values):
            header_row = row
            break
    if header_row is None:
        return pd.DataFrame(columns=EVIDENCE_COLUMNS), match
    return _normalize_evidence_columns(_read_rectangular_table(ws, header_row)), match


@st.cache_data(show_spinner=False)
def parse_workbook(workbook_bytes: bytes) -> dict[str, Any]:
    """Parse the workbook into compact, reusable dashboard tables."""
    wb = load_workbook(BytesIO(workbook_bytes), data_only=True, read_only=False)
    missing = sorted(REQUIRED_SHEETS.difference(wb.sheetnames))
    if missing:
        raise ValueError("Missing required sheet(s): " + ", ".join(missing))

    level_ws = wb["Level1_AHP"]
    framework_header = _locate_header(level_ws, {"Pillar", "Code", "Indicator", "Global Weight"})
    framework = _read_rectangular_table(level_ws, framework_header)
    framework = framework[["Pillar", "Code", "Indicator", "Local Weight", "Pillar Weight", "Global Weight"]].copy()
    framework = framework[framework["Code"].astype(str).str.match(r"^[ESG]\d+$", na=False)]
    for column in ["Local Weight", "Pillar Weight", "Global Weight"]:
        framework[column] = pd.to_numeric(framework[column], errors="coerce")
    if framework["Global Weight"].isna().any() or framework["Global Weight"].sum() <= 0:
        raise ValueError("The final AHP framework has missing or invalid global weights.")
    framework["Global Weight"] = framework["Global Weight"] / framework["Global Weight"].sum()
    framework["Code"] = framework["Code"].astype(str).str.strip()
    framework["Pillar"] = framework["Pillar"].astype(str).str.strip()
    framework["Indicator"] = framework["Indicator"].astype(str).str.strip()
    framework = framework.reset_index(drop=True)

    rubric_ws = wb["Rubric"]
    rubric_header = _locate_header(rubric_ws, {"Pillar", "Code", "Indicator", "Evidence to record"})
    rubric = _read_rectangular_table(rubric_ws, rubric_header)
    rubric = rubric[rubric["Code"].astype(str).str.match(r"^[ESG]\d+$", na=False)].reset_index(drop=True)

    topsis_ws = wb["TOPSIS"]
    score_header = _locate_header(topsis_ws, {"Vendor", "Evidence Status"})
    code_columns = framework["Code"].tolist()
    scores: list[dict[str, Any]] = []
    for row in range(score_header + 1, topsis_ws.max_row + 1):
        vendor = topsis_ws.cell(row, 1).value
        if vendor is None:
            if scores:
                break
            continue
        if str(vendor).strip().lower() in {"vector length", "vendor"}:
            break
        record: dict[str, Any] = {"Supplier": str(vendor).strip()}
        for offset, code in enumerate(code_columns, start=2):
            record[code] = topsis_ws.cell(row, offset).value
        scores.append(record)
    score_df = pd.DataFrame(scores, columns=["Supplier", *code_columns])
    for code in code_columns:
        score_df[code] = pd.to_numeric(score_df[code], errors="coerce")

    consistency_rows: list[dict[str, Any]] = []
    matrix_names = {
        "Level1_AHP": "ESG Pillars",
        "E_AHP": "Environment",
        "S_AHP": "Social",
        "G_AHP": "Governance",
    }
    rationales: list[dict[str, str]] = []
    for sheet_name, matrix_name in matrix_names.items():
        ws = wb[sheet_name]
        cr = pd.to_numeric(_cell_below_label(ws, "CR"), errors="coerce")
        status = _cell_below_label(ws, "Status")
        consistency_rows.append(
            {
                "Matrix": matrix_name,
                "Consistency Ratio": float(cr) if pd.notna(cr) else np.nan,
                "Threshold": 0.10,
                "Status": str(status).strip() if status is not None else "Unknown",
            }
        )
        rationale = _cell_below_label(ws, "Judgment Rationale")
        if rationale:
            rationales.append({"Matrix": matrix_name, "Rationale": str(rationale)})

    evidence, evidence_sheet = _parse_evidence_sheet(wb)
    pillar_weights = (
        framework.groupby("Pillar", as_index=False)["Global Weight"]
        .sum()
        .rename(columns={"Global Weight": "Weight"})
    )
    return {
        "framework": framework,
        "rubric": rubric,
        "scores": score_df,
        "pillar_weights": pillar_weights,
        "consistency": pd.DataFrame(consistency_rows),
        "rationales": pd.DataFrame(rationales),
        "evidence": evidence,
        "evidence_sheet": evidence_sheet,
        "sheet_names": wb.sheetnames,
    }


def sanitize_scores(raw: pd.DataFrame, codes: list[str]) -> tuple[pd.DataFrame, list[str]]:
    """Validate session edits while preserving missing values."""
    scores = raw.copy()
    messages: list[str] = []
    if "Supplier" not in scores.columns:
        scores.insert(0, "Supplier", "")
    scores["Supplier"] = scores["Supplier"].fillna("").astype(str).str.strip()
    scores = scores[scores["Supplier"] != ""].copy()
    for code in codes:
        if code not in scores.columns:
            scores[code] = np.nan
        scores[code] = pd.to_numeric(scores[code], errors="coerce")
        invalid = scores[code].notna() & ~scores[code].between(0, 5)
        if invalid.any():
            scores.loc[invalid, code] = np.nan
            messages.append(f"{code}: values outside 0–5 were cleared.")
    duplicated = scores["Supplier"].duplicated(keep=False)
    if duplicated.any():
        names = ", ".join(sorted(scores.loc[duplicated, "Supplier"].unique()))
        messages.append(f"Duplicate supplier names found: {names}.")
    return scores[["Supplier", *codes]].reset_index(drop=True), messages


def calculate_topsis(
    scores: pd.DataFrame, framework: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Calculate benefit-type TOPSIS for suppliers with all 12 scores."""
    codes = framework["Code"].tolist()
    weights = framework.set_index("Code").loc[codes, "Global Weight"].to_numpy(dtype=float)
    complete_mask = scores[codes].notna().all(axis=1)
    complete = scores.loc[complete_mask, ["Supplier", *codes]].copy()
    if len(complete) < 2:
        return pd.DataFrame(), pd.DataFrame(), complete

    matrix = complete[codes].to_numpy(dtype=float)
    vector_lengths = np.sqrt(np.square(matrix).sum(axis=0))
    normalized = np.divide(matrix, vector_lengths, out=np.zeros_like(matrix), where=vector_lengths != 0)
    weighted = normalized * weights
    ideal_best = weighted.max(axis=0)
    ideal_worst = weighted.min(axis=0)
    distance_best = np.sqrt(np.square(weighted - ideal_best).sum(axis=1))
    distance_worst = np.sqrt(np.square(weighted - ideal_worst).sum(axis=1))
    denominators = distance_best + distance_worst
    closeness = np.divide(
        distance_worst,
        denominators,
        out=np.full(distance_worst.shape, 0.5, dtype=float),
        where=denominators != 0,
    )
    ranking = pd.DataFrame(
        {
            "Supplier": complete["Supplier"].to_numpy(),
            "Distance to Best (D+)": distance_best,
            "Distance to Worst (D-)": distance_worst,
            "Closeness Coefficient": closeness,
        }
    ).sort_values(["Closeness Coefficient", "Supplier"], ascending=[False, True])
    ranking["Rank"] = ranking["Closeness Coefficient"].rank(method="min", ascending=False).astype(int)
    ranking = ranking.sort_values(["Rank", "Supplier"]).reset_index(drop=True)
    weighted_df = pd.DataFrame(weighted, columns=codes)
    weighted_df.insert(0, "Supplier", complete["Supplier"].to_numpy())
    return ranking, weighted_df, complete


def score_completeness(scores: pd.DataFrame, codes: list[str]) -> pd.DataFrame:
    result = scores[["Supplier"]].copy()
    result["Scored Indicators"] = scores[codes].notna().sum(axis=1)
    result["Total Indicators"] = len(codes)
    result["Completion"] = result["Scored Indicators"] / len(codes)
    result["Status"] = np.where(result["Completion"].eq(1), "Complete", "Incomplete")
    return result


def pillar_scores(scores: pd.DataFrame, framework: pd.DataFrame) -> pd.DataFrame:
    """Calculate within-pillar weighted averages on the original 0–5 scale."""
    rows: list[dict[str, Any]] = []
    for _, supplier_row in scores.iterrows():
        for pillar, group in framework.groupby("Pillar", sort=False):
            codes = group["Code"].tolist()
            values = pd.to_numeric(supplier_row[codes], errors="coerce")
            available = values.notna().to_numpy()
            weights = group["Global Weight"].to_numpy(dtype=float)
            value_array = values.to_numpy(dtype=float)
            score = np.average(value_array[available], weights=weights[available]) if available.any() else np.nan
            rows.append({"Supplier": supplier_row["Supplier"], "Pillar": pillar, "Pillar Score": score})
    return pd.DataFrame(rows)


def format_percent(value: float) -> str:
    return f"{value:.1%}" if pd.notna(value) else "—"


def base_chart_layout(fig: go.Figure, height: int = 420) -> go.Figure:
    fig.update_layout(
        height=height,
        margin=dict(l=16, r=16, t=52, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend_title_text="",
        font=dict(family="Arial, sans-serif"),
    )
    return fig


def render_overview(
    scores: pd.DataFrame,
    framework: pd.DataFrame,
    pillar_weight_df: pd.DataFrame,
    consistency: pd.DataFrame,
    ranking: pd.DataFrame,
    evidence: pd.DataFrame,
) -> None:
    codes = framework["Code"].tolist()
    completion = score_completeness(scores, codes)
    complete_count = int(completion["Status"].eq("Complete").sum())
    consistency_pass = int(consistency["Status"].str.upper().eq("PASS").sum())
    top_supplier = ranking.iloc[0]["Supplier"] if not ranking.empty else "Awaiting scores"
    top_score = ranking.iloc[0]["Closeness Coefficient"] if not ranking.empty else np.nan

    st.subheader("Executive Overview")
    st.caption("Decision snapshot for procurement and sustainability reviewers")
    cols = st.columns(5)
    cols[0].metric("Suppliers", len(scores))
    cols[1].metric("Complete assessments", f"{complete_count}/{len(scores)}")
    cols[2].metric("Top supplier", top_supplier)
    cols[3].metric("TOPSIS closeness", f"{top_score:.3f}" if pd.notna(top_score) else "—")
    cols[4].metric("AHP matrices passed", f"{consistency_pass}/{len(consistency)}")

    left, right = st.columns([1, 1.25])
    with left:
        pillar_plot = pillar_weight_df.copy()
        fig = px.bar(
            pillar_plot,
            x="Weight",
            y="Pillar",
            orientation="h",
            color="Pillar",
            color_discrete_map=PILLAR_COLORS,
            text=pillar_plot["Weight"].map(lambda value: f"{value:.1%}"),
            title="AHP pillar priorities",
        )
        fig.update_xaxes(tickformat=".0%", range=[0, max(0.6, pillar_plot["Weight"].max() * 1.15)], title=None)
        fig.update_yaxes(title=None, categoryorder="array", categoryarray=pillar_plot["Pillar"].tolist()[::-1])
        fig.update_traces(textposition="outside", hovertemplate="%{y}: %{x:.2%}<extra></extra>")
        fig.update_layout(showlegend=False)
        st.plotly_chart(base_chart_layout(fig, 350), width="stretch")
    with right:
        completion_plot = completion.sort_values("Completion")
        fig = px.bar(
            completion_plot,
            x="Completion",
            y="Supplier",
            orientation="h",
            color="Status",
            color_discrete_map=STATUS_COLORS,
            text=completion_plot.apply(lambda row: f"{int(row['Scored Indicators'])}/{int(row['Total Indicators'])}", axis=1),
            title="Assessment completeness",
        )
        fig.update_xaxes(tickformat=".0%", range=[0, 1.05], title=None)
        fig.update_yaxes(title=None)
        fig.update_traces(textposition="inside", hovertemplate="%{y}: %{x:.0%}<extra></extra>")
        fig.update_layout(legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
        st.plotly_chart(base_chart_layout(fig, 350), width="stretch")

    if ranking.empty:
        st.info("TOPSIS ranking will appear after at least two suppliers have complete 0–5 scores for all 12 indicators.")
    else:
        st.markdown("#### Current ranking")
        display = ranking[["Rank", "Supplier", "Closeness Coefficient"]].copy()
        st.dataframe(
            display.style.format({"Closeness Coefficient": "{:.3f}"}),
            width="stretch",
            hide_index=True,
        )

    if evidence.empty:
        st.caption("Evidence Register: not connected yet. The Evidence Explorer tab shows the supported import schema.")
    else:
        covered = evidence.dropna(subset=["Supplier", "Code"]).drop_duplicates(["Supplier", "Code"])
        possible = max(len(scores) * len(codes), 1)
        st.caption(f"Evidence Register: {len(evidence)} records loaded; indicator coverage {len(covered)}/{possible} ({len(covered) / possible:.1%}).")


def render_comparison(scores: pd.DataFrame, framework: pd.DataFrame) -> None:
    st.subheader("Supplier Comparison")
    st.caption("Compare raw rubric scores and weighted ESG pillar performance")
    if scores.empty:
        st.warning("No suppliers were found in the TOPSIS input table.")
        return

    supplier_options = scores["Supplier"].tolist()
    defaults = supplier_options[: min(3, len(supplier_options))]
    selected = st.multiselect("Suppliers to compare", supplier_options, default=defaults, max_selections=5)
    if not selected:
        st.info("Select at least one supplier.")
        return
    selected_scores = scores[scores["Supplier"].isin(selected)].copy()
    codes = framework["Code"].tolist()

    heatmap_values = selected_scores.set_index("Supplier")[codes]
    hover_text = np.where(heatmap_values.notna(), heatmap_values.round(1).astype(str), "Not scored")
    fig = go.Figure(
        data=go.Heatmap(
            z=heatmap_values.to_numpy(dtype=float),
            x=codes,
            y=heatmap_values.index.tolist(),
            zmin=0,
            zmax=5,
            colorscale=[[0, "#F1F4F8"], [0.5, "#8BC8B7"], [1, "#08765A"]],
            text=hover_text,
            hovertemplate="Supplier: %{y}<br>Indicator: %{x}<br>Score: %{text}<extra></extra>",
            colorbar=dict(title="Score"),
        )
    )
    fig.update_layout(title="Indicator score heatmap")
    st.plotly_chart(base_chart_layout(fig, 300 + 35 * len(selected)), width="stretch")

    left, right = st.columns([1.05, 1])
    with left:
        radar = go.Figure()
        radar_categories = codes + [codes[0]]
        for _, row in selected_scores.iterrows():
            values = [row[code] if pd.notna(row[code]) else None for code in codes]
            radar.add_trace(
                go.Scatterpolar(
                    r=values + [values[0]],
                    theta=radar_categories,
                    fill="toself",
                    opacity=0.55,
                    name=row["Supplier"],
                    connectgaps=False,
                )
            )
        radar.update_layout(
            title="ESG indicator profile",
            polar=dict(radialaxis=dict(visible=True, range=[0, 5], dtick=1)),
        )
        st.plotly_chart(base_chart_layout(radar, 470), width="stretch")
    with right:
        pillar = pillar_scores(selected_scores, framework).dropna(subset=["Pillar Score"])
        if pillar.empty:
            st.info("Pillar scores will appear when scores are entered.")
        else:
            fig = px.bar(
                pillar,
                x="Pillar",
                y="Pillar Score",
                color="Supplier",
                barmode="group",
                text_auto=".2f",
                title="Within-pillar weighted scores",
            )
            fig.update_yaxes(range=[0, 5], dtick=1, title="Score (0–5)")
            fig.update_xaxes(title=None)
            st.plotly_chart(base_chart_layout(fig, 470), width="stretch")

    labels = framework.set_index("Code")["Indicator"].to_dict()
    detail = selected_scores.set_index("Supplier")[codes].T
    detail.insert(0, "Indicator", [labels[code] for code in detail.index])
    detail.insert(0, "Code", detail.index)
    st.markdown("#### Score detail")
    st.dataframe(detail.reset_index(drop=True), width="stretch", hide_index=True)


def render_ahp(
    framework: pd.DataFrame,
    pillar_weight_df: pd.DataFrame,
    consistency: pd.DataFrame,
    rationales: pd.DataFrame,
) -> None:
    st.subheader("AHP Weights")
    st.caption("Two-level AHP priorities and pairwise-comparison consistency")
    left, right = st.columns([0.8, 1.4])
    with left:
        fig = px.pie(
            pillar_weight_df,
            values="Weight",
            names="Pillar",
            color="Pillar",
            color_discrete_map=PILLAR_COLORS,
            hole=0.62,
            title="Pillar weight allocation",
        )
        fig.update_traces(textinfo="label+percent", hovertemplate="%{label}: %{value:.2%}<extra></extra>")
        st.plotly_chart(base_chart_layout(fig, 430), width="stretch")
    with right:
        indicator_plot = framework.sort_values("Global Weight", ascending=True).copy()
        indicator_plot["Label"] = indicator_plot["Code"] + " · " + indicator_plot["Indicator"]
        fig = px.bar(
            indicator_plot,
            x="Global Weight",
            y="Label",
            orientation="h",
            color="Pillar",
            color_discrete_map=PILLAR_COLORS,
            text=indicator_plot["Global Weight"].map(lambda value: f"{value:.1%}"),
            title="Global indicator weights",
        )
        fig.update_xaxes(tickformat=".0%", title=None)
        fig.update_yaxes(title=None)
        fig.update_traces(textposition="outside", hovertemplate="%{y}<br>Weight: %{x:.2%}<extra></extra>")
        st.plotly_chart(base_chart_layout(fig, 520), width="stretch")

    st.markdown("#### Consistency checks")
    consistency_display = consistency.copy()
    consistency_display["CR < 0.10"] = consistency_display["Consistency Ratio"].lt(0.10)
    st.dataframe(
        consistency_display.style.format({"Consistency Ratio": "{:.4f}", "Threshold": "{:.2f}"}),
        width="stretch",
        hide_index=True,
    )
    if not rationales.empty:
        with st.expander("Judgment rationales from the workbook"):
            for _, row in rationales.iterrows():
                st.markdown(f"**{row['Matrix']}** — {row['Rationale']}")

    st.markdown("#### Weight table")
    weight_table = framework.copy()
    st.dataframe(
        weight_table.style.format(
            {"Local Weight": "{:.2%}", "Pillar Weight": "{:.2%}", "Global Weight": "{:.2%}"}
        ),
        width="stretch",
        hide_index=True,
    )


def render_topsis(scores: pd.DataFrame, framework: pd.DataFrame, ranking: pd.DataFrame) -> None:
    st.subheader("TOPSIS Ranking")
    st.caption("Higher closeness to the positive ideal solution indicates stronger overall ESG performance")
    codes = framework["Code"].tolist()
    completeness = score_completeness(scores, codes)
    if ranking.empty:
        st.info("At least two fully scored suppliers are required for a meaningful TOPSIS ranking.")
        st.dataframe(completeness, width="stretch", hide_index=True)
    else:
        ranking_plot = ranking.sort_values("Closeness Coefficient", ascending=True)
        fig = px.bar(
            ranking_plot,
            x="Closeness Coefficient",
            y="Supplier",
            orientation="h",
            color="Closeness Coefficient",
            color_continuous_scale=["#D6E3DF", "#159A74"],
            text=ranking_plot["Closeness Coefficient"].map(lambda value: f"{value:.3f}"),
            title="Supplier closeness to the ideal solution",
        )
        fig.update_xaxes(range=[0, 1], title="Closeness coefficient")
        fig.update_yaxes(title=None)
        fig.update_traces(textposition="outside")
        fig.update_layout(coloraxis_showscale=False)
        st.plotly_chart(base_chart_layout(fig, 330 + 35 * len(ranking)), width="stretch")

        display = ranking[["Rank", "Supplier", "Closeness Coefficient", "Distance to Best (D+)", "Distance to Worst (D-)"]]
        st.dataframe(
            display.style.format(
                {
                    "Closeness Coefficient": "{:.4f}",
                    "Distance to Best (D+)": "{:.4f}",
                    "Distance to Worst (D-)": "{:.4f}",
                }
            ),
            width="stretch",
            hide_index=True,
        )

    with st.expander("Method and inclusion rule"):
        st.markdown(
            "All 12 indicators are treated as benefit criteria. The app uses vector normalization, "
            "multiplies by the workbook's AHP global weights, calculates Euclidean distance to the "
            "positive and negative ideal solutions, and ranks by the closeness coefficient. Suppliers "
            "with any missing indicator score remain visible in completeness reporting but are excluded "
            "from ranking to avoid presenting partial assessments as comparable results."
        )


def render_evidence_explorer(
    evidence: pd.DataFrame,
    framework: pd.DataFrame,
    rubric: pd.DataFrame,
    evidence_sheet: str | None,
) -> None:
    st.subheader("Evidence Explorer")
    st.caption("Optional data interface for traceable Supplier → Indicator → Evidence → Score records")
    if evidence.empty:
        if evidence_sheet:
            st.warning(f"Sheet '{evidence_sheet}' was found, but no recognizable Evidence Register records could be read.")
        else:
            st.info("No Evidence Register sheet or CSV is connected. Add a sheet named 'Evidence_Register' or upload a CSV from the sidebar.")
        template = pd.DataFrame(columns=EVIDENCE_COLUMNS)
        st.download_button(
            "Download Evidence Register CSV template",
            data=template.to_csv(index=False).encode("utf-8-sig"),
            file_name="Evidence_Register_Template.csv",
            mime="text/csv",
        )
        st.markdown("#### Expected fields")
        st.dataframe(pd.DataFrame({"Field": EVIDENCE_COLUMNS}), width="stretch", hide_index=True)
    else:
        col1, col2, col3 = st.columns(3)
        suppliers = sorted(evidence["Supplier"].dropna().astype(str).unique())
        codes = framework["Code"].tolist()
        statuses = sorted(evidence["Verification Status"].dropna().astype(str).unique())
        supplier_filter = col1.multiselect("Supplier", suppliers)
        code_filter = col2.multiselect("Indicator code", codes)
        status_filter = col3.multiselect("Verification status", statuses)
        filtered = evidence.copy()
        if supplier_filter:
            filtered = filtered[filtered["Supplier"].astype(str).isin(supplier_filter)]
        if code_filter:
            filtered = filtered[filtered["Code"].astype(str).isin(code_filter)]
        if status_filter:
            filtered = filtered[filtered["Verification Status"].astype(str).isin(status_filter)]
        metrics = st.columns(3)
        metrics[0].metric("Evidence records", len(filtered))
        metrics[1].metric("Suppliers covered", filtered["Supplier"].nunique())
        metrics[2].metric("Indicators covered", filtered["Code"].nunique())
        st.dataframe(filtered, width="stretch", hide_index=True)
        st.download_button(
            "Download filtered evidence",
            data=filtered.to_csv(index=False).encode("utf-8-sig"),
            file_name="Evidence_Register_Filtered.csv",
            mime="text/csv",
        )

    with st.expander("Rubric evidence requirements"):
        columns = [column for column in ["Pillar", "Code", "Indicator", "Evidence to record"] if column in rubric.columns]
        st.dataframe(rubric[columns], width="stretch", hide_index=True)


def main() -> None:
    st.title("ESG Supplier Sustainability Dashboard")
    st.markdown(
        "<div class='source-note'>Evidence-based 0–5 scoring · two-level AHP weights · benefit-type TOPSIS ranking</div>",
        unsafe_allow_html=True,
    )

    default_path = _find_default_workbook()
    st.sidebar.header("Data source")
    uploaded_workbook = st.sidebar.file_uploader("Use another AHP–TOPSIS workbook", type=["xlsx"])
    if uploaded_workbook is not None:
        workbook_bytes = uploaded_workbook.getvalue()
        source_label = uploaded_workbook.name
    elif default_path is not None:
        workbook_bytes = default_path.read_bytes()
        source_label = str(default_path)
    else:
        st.warning("No model workbook was found. Upload an .xlsx file in the sidebar or set ESG_WORKBOOK_PATH.")
        st.stop()

    try:
        data = parse_workbook(workbook_bytes)
    except Exception as exc:
        st.error(f"The workbook could not be loaded: {exc}")
        st.stop()

    st.sidebar.caption(f"Workbook: {source_label}")
    st.sidebar.caption("Required sheets: ✓ " + ", ".join(sorted(REQUIRED_SHEETS)))

    evidence_upload = st.sidebar.file_uploader("Optional Evidence Register CSV", type=["csv"])
    evidence = data["evidence"]
    if evidence_upload is not None:
        try:
            evidence = _normalize_evidence_columns(pd.read_csv(evidence_upload))
        except Exception as exc:
            st.sidebar.error(f"Evidence CSV could not be read: {exc}")

    framework = data["framework"]
    codes = framework["Code"].tolist()
    original_scores = data["scores"]
    edit_scores = st.sidebar.toggle("Edit scores in this session", value=original_scores[codes].isna().any().any())
    scores = original_scores.copy()
    if edit_scores:
        with st.expander("Working score inputs (session only)", expanded=original_scores[codes].isna().any().any()):
            st.caption("Enter ordinal scores from 0 to 5. Changes update the dashboard but do not write back to the workbook.")
            editor_key = "score_editor_" + str(abs(hash(workbook_bytes[:2048])))
            edited = st.data_editor(
                original_scores,
                key=editor_key,
                num_rows="dynamic",
                width="stretch",
                hide_index=True,
                column_config={
                    "Supplier": st.column_config.TextColumn("Supplier", required=True),
                    **{
                        code: st.column_config.NumberColumn(code, min_value=0, max_value=5, step=1, format="%d")
                        for code in codes
                    },
                },
            )
            scores, validation_messages = sanitize_scores(edited, codes)
            for message in validation_messages:
                st.warning(message)
    else:
        scores, _ = sanitize_scores(scores, codes)

    ranking, _, _ = calculate_topsis(scores, framework)

    tabs = st.tabs(
        [
            "Executive Overview",
            "Supplier Comparison",
            "AHP Weights",
            "TOPSIS Ranking",
            "Evidence Explorer",
        ]
    )
    with tabs[0]:
        render_overview(scores, framework, data["pillar_weights"], data["consistency"], ranking, evidence)
    with tabs[1]:
        render_comparison(scores, framework)
    with tabs[2]:
        render_ahp(framework, data["pillar_weights"], data["consistency"], data["rationales"])
    with tabs[3]:
        render_topsis(scores, framework, ranking)
    with tabs[4]:
        render_evidence_explorer(evidence, framework, data["rubric"], data["evidence_sheet"])

    st.divider()
    st.caption(f"Source workbook: {source_label} · Scores and weights are interpreted from the workbook; session edits are not persisted.")


if __name__ == "__main__":
    main()
